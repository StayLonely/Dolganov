from openpyxl import Workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.styles import Font, Border, Side


def updateCell(cell,t,text):
    ExcelSheet[cell] = text
    ExcelSheet[cell].border = Border(top=t, left=t, right=t, bottom=t)
    ExcelSheet[cell].font = Font(bold=True)

def updateRow(text, name, r):
    ExcelSheet[f"{text}{r}"] = name
    ExcelSheet[f"{text}{r}"].border = Border(top=SheetStyle, left=SheetStyle, right=SheetStyle, bottom=SheetStyle)

def sizeTable():
    for column_cells in ExcelSheet.columns:
        arr = []
        for cell in column_cells:
            arr.append(len(textControl(cell.value)))
        length = max(arr) + 2
        ExcelSheet.column_dimensions[column_cells[0].column_letter].width = length

def printStat():
    print("Динамика уровня зарплат по годам: " + str(DateBaseSalaryYears))
    print("Динамика количества вакансий по годам: " + str(DateBaseYearsCount))
    print("Динамика уровня зарплат по годам для выбранной профессии: " + str(DateBaseSalaryV))
    print("Динамика количества вакансий по годам для выбранной профессии: " + str(DateBaseCountV))
    print("Уровень зарплат по городам (в порядке убывания): " + str(DateBaseSalaryAr))
    print("Доля вакансий по городам (в порядке убывания): " + str(DateBaseCountAr))



def textControl(value):
    return "" if value is None else str(value)



DateBaseSalaryYears = {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510,
                       2014: 50658, 2015: 52696, 2016: 62675, 2017: 60935, 2018: 58335, 2019: 69467, 2020: 73431,
                       2021: 82690, 2022: 91795}
DateBaseYearsCount = {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954,
                      2014: 66837, 2015: 70039, 2016: 75145, 2017: 82823, 2018: 131701, 2019: 115086, 2020: 102243,
                      2021: 57623, 2022: 18294}
DateBaseSalaryV = {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960,
                   2013: 58804, 2014: 62384, 2015: 62322, 2016: 66817, 2017: 72460, 2018: 76879,
                   2019: 85300, 2020: 89791, 2021: 100987, 2022: 116651}
DateBaseCountV = {2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990,
                  2014: 5492, 2015: 5375, 2016: 7219, 2017: 8105, 2018: 10062, 2019: 9016, 2020: 7113,
                  2021: 3466, 2022: 1115}
DateBaseSalaryAr = {'Москва': 76970, 'Санкт-Петербург': 65286, 'Новосибирск': 62254, 'Екатеринбург': 60962,
                           'Казань': 52580, 'Краснодар': 51644, 'Челябинск': 51265, 'Самара': 50994, 'Пермь': 48089,
                           'Нижний Новгород': 47662}
DateBaseCountAr = {'Москва': 0.3246, 'Санкт-Петербург': 0.1197, 'Новосибирск': 0.0271, 'Казань': 0.0237,
                          'Нижний Новгород': 0.0232, 'Ростов-на-Дону': 0.0209, 'Екатеринбург': 0.0207,
                          'Краснодар': 0.0185, 'Самара': 0.0143, 'Воронеж': 0.0141}

DateBaseSalaryYears = {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510, 2014: 50658}
DateBaseYearsCount = {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954, 2014: 66837}
DateBaseSalaryV = {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960, 2013: 58804, 2014: 62384}
DateBaseCountV = {2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990, 2014: 5492}
DateBaseSalaryAr = {'Москва': 57354, 'Санкт-Петербург': 46291, 'Новосибирск': 41580, 'Екатеринбург': 41091, 'Казань': 37587, 'Самара': 34091, 'Нижний Новгород': 33637, 'Ярославль': 32744, 'Краснодар': 32542, 'Воронеж': 29725}
DateBaseCountAr = {'Москва': 0.4581, 'Санкт-Петербург': 0.1415, 'Нижний Новгород': 0.0269, 'Казань': 0.0266, 'Ростов-на-Дону': 0.0234, 'Новосибирск': 0.0202, 'Екатеринбург': 0.0143, 'Воронеж': 0.014, 'Самара': 0.0133, 'Краснодар': 0.0131}

myExcel = Workbook()
del myExcel['Sheet']
ExcelSheet = myExcel.create_sheet('Статистика по годам')
SheetStyle = Side(border_style="thin", color="000000")
updateCell("A1", SheetStyle, "Год")
updateCell("B1", SheetStyle, "Средняя зарплата")
updateCell("C1", SheetStyle, "Средняя зарплата - Программист")
updateCell("D1", SheetStyle, "Количество вакансий")
updateCell("E1", SheetStyle, "Количество вакансий - Программист")

for row, (year, value) in enumerate(DateBaseSalaryYears.items(), start=2):
    updateRow('A', year, row)
    updateRow('B', value, row)
    updateRow('C', DateBaseSalaryV[year], row)
    updateRow('D', DateBaseYearsCount[year], row)
    updateRow('E', DateBaseCountV[year], row)

sizeTable()

ExcelSheet = myExcel.create_sheet('Статистика по городам')
updateCell("A1", SheetStyle, "Город")
updateCell("B1", SheetStyle, "Уровень зарплат")
updateCell("D1", SheetStyle, "Город")
updateCell("E1", SheetStyle, "Доля вакансий")

for row, (year, value) in enumerate(DateBaseSalaryAr.items(), start=2):
    updateRow('A', year, row)
    updateRow('B', value, row)
for row, (year, value) in enumerate(DateBaseCountAr.items(), start=2):
    updateRow('D', year, row)
    updateRow('E',  value, row)
    ExcelSheet[f"E{row}"].number_format = FORMAT_PERCENTAGE_00

sizeTable()

printStat()

myExcel.save('report.xlsx')